import os
import openpyxl
import pandas as pd
import json

base_dir = r"c:\Users\computer1\Desktop\Growthik_Media\02_Clients\03_SearchSarkariNaukri\SearchSarkariNaukri\.agents\06_Keywords"
excel_path = r"C:\Users\computer1\Desktop\Growthik_Media\02_Clients\03_SearchSarkariNaukri\SearchSarkariNaukri\SearchSarkariNaukri-Keyword-List-899.xlsx"
excel_backup_path = r"C:\Users\computer1\Desktop\Growthik_Media\02_Clients\03_SearchSarkariNaukri\SearchSarkariNaukri\SearchSarkariNaukri-Keyword-Master-Updated.xlsx"

os.makedirs(base_dir, exist_ok=True)

print("Starting Keyword Universe Master Processor...")

# 1. Prompt Keyword Universes Definition
prompt_clusters = {
    "01_General_Sarkari_Naukri": {
        "title": "A. General Sarkari Naukri",
        "category": "Core Head Terms",
        "sub_category": "Head & General Intent",
        "intent": "Navigational/Commercial",
        "target_page": "/ or /jobs",
        "keywords": [
            # Google Trends Top Queries (General Govt Jobs)
            "government jobs 2026", "government jobs india", "government jobs recruitment", "central government jobs", "government jobs vacancy",
            # Google Trends Rising Queries
            "ugc net", "government jobs in goa", "post office recruitment 2026", "government part time jobs work from home",
            "central government jobs in karnataka", "hr government jobs", "jharkhand government jobs vacancy 2026", "aai recruitment 2026",
            # Original Head & General Keywords
            "sarkari naukri", "government jobs", "govt jobs 2026", "latest government jobs", "latest govt jobs",
            "government jobs today", "govt jobs today", "sarkari job", "sarkari jobs", "सरकारी नोकरी", "सरकारी नौकरी 2026", "सरकारी जॉब"
        ]
    },
    "02_Latest_Jobs": {
        "title": "B. Latest Jobs Universe",
        "category": "Core Head Terms",
        "sub_category": "Freshness & New Vacancies",
        "intent": "Commercial/Informational",
        "target_page": "/jobs or /job-updates",
        "keywords": [
            "latest sarkari naukri", "latest government jobs", "latest govt jobs", "new government jobs",
            "new govt jobs 2026", "latest recruitment 2026", "government recruitment 2026", "upcoming government jobs",
            "upcoming govt jobs", "current government jobs", "active government jobs", "government vacancies 2026",
            "govt vacancies 2026"
        ]
    },
    "03_Deadline_Closing_Soon": {
        "title": "C. Deadline & Closing Soon Keywords",
        "category": "Informational/Long-tail",
        "sub_category": "Urgency & Closing Dates",
        "intent": "Transactional/Commercial",
        "target_page": "/jobs (Filter: Last Date Ending Soon)",
        "keywords": [
            # Google Trends Top & Rising Deadline Queries
            "mpsc last date", "talathi bharti last date 2026", "talathi bharti 2026 apply online last date",
            # Original Deadline Keywords
            "government jobs last date", "govt jobs last date", "sarkari naukri last date",
            "government jobs closing soon", "government jobs ending today", "government jobs apply today",
            "government jobs last date today", "latest government jobs last date", "govt jobs apply online",
            "government job apply online"
        ]
    },
    "04_Qualification_Universe": {
        "title": "Qualification Keyword Universe",
        "category": "Qualification-wise",
        "sub_category": "Education Filter Pages",
        "intent": "Commercial/Informational",
        "target_page": "/qualifications/{10th|12th|iti|diploma|graduate|engineering}",
        "keywords": [
            "10th pass government jobs", "10th pass govt jobs", "10th pass sarkari naukri", "government jobs after 10th",
            "govt jobs after 10th", "10th pass government jobs 2026", "10th pass govt jobs 2026", "10th pass railway jobs",
            "10th pass police jobs", "10th pass government jobs Maharashtra",
            "12th pass government jobs", "12th pass govt jobs", "12th pass sarkari naukri", "government jobs after 12th",
            "govt jobs after 12th", "12th pass government jobs 2026", "12th pass railway jobs", "12th pass police jobs",
            "12th pass jobs Maharashtra",
            "graduate government jobs", "graduate govt jobs", "government jobs for graduates", "graduate government jobs 2026",
            "graduate sarkari naukri", "government jobs after graduation", "graduate jobs without experience",
            "ITI government jobs", "ITI govt jobs", "ITI jobs 2026", "government jobs for ITI", "ITI apprentice jobs",
            "ITI railway jobs", "ITI government jobs Maharashtra",
            "diploma government jobs", "diploma govt jobs", "diploma jobs 2026", "government jobs after diploma",
            "diploma railway jobs", "diploma government jobs Maharashtra",
            "engineering government jobs", "government jobs for engineers", "BTech government jobs", "BE government jobs",
            "engineering govt jobs 2026", "BTech government jobs 2026"
        ]
    },
    "05_Maharashtra_Districts_Universe": {
        "title": "Maharashtra & Local District Keyword Universe",
        "category": "District-wise",
        "sub_category": "Maharashtra & City/District Pages",
        "intent": "Local Commercial/Informational",
        "target_page": "/districts/{district-name} or /category/state-government-jobs",
        "keywords": [
            # Google Trends Rising Local/State Queries
            "mumbai police bharti", "government jobs in goa", "central government jobs in karnataka", "jharkhand government jobs vacancy 2026",
            "digital satbara", "digital 712", "mahabhumi", "rcms", "mahakosh", "caste validity", "mahajyoti",
            # Original Maharashtra & District Keywords
            "Maharashtra government jobs", "Maharashtra govt jobs", "Maharashtra government jobs 2026", "Maharashtra sarkari naukri",
            "Maharashtra govt recruitment", "Maharashtra recruitment 2026", "Maharashtra government vacancy", "Maharashtra government jobs today",
            "government jobs in Pune", "govt jobs in Pune", "Pune government jobs 2026", "Pune govt recruitment", "Pune sarkari naukri", "government jobs Pune", "Pune government vacancy",
            "government jobs in Mumbai", "govt jobs Mumbai", "Mumbai government jobs 2026", "Mumbai govt recruitment", "Mumbai sarkari naukri",
            "Nagpur government jobs", "govt jobs Nagpur", "Nagpur sarkari naukri",
            "Nashik government jobs", "govt jobs Nashik", "Nashik sarkari naukri",
            "Thane government jobs", "govt jobs Thane", "Thane sarkari naukri",
            "Navi Mumbai government jobs", "govt jobs Navi Mumbai",
            "Solapur government jobs", "govt jobs Solapur",
            "Kolhapur government jobs", "govt jobs Kolhapur",
            "Latur government jobs", "govt jobs Latur",
            "Beed government jobs", "govt jobs Beed",
            "Dharashiv government jobs", "govt jobs Dharashiv",
            "Amravati government jobs", "govt jobs Amravati",
            "Chhatrapati Sambhajinagar government jobs", "govt jobs Sambhajinagar", "Chhatrapati Sambhajinagar sarkari naukri"
        ]
    },
    "06_MPSC_Talathi_GroupC_Universe": {
        "title": "MPSC, Talathi & Group C Keyword Universe",
        "category": "Department/Exam-wise",
        "sub_category": "MPSC & State Level Exams",
        "intent": "Informational/Commercial",
        "target_page": "/jobs?category=mpsc or /jobs?category=talathi",
        "keywords": [
            # Google Trends Top Queries (MPSC & Talathi)
            "mpsc login", "group c mpsc", "mpsc group c", "mpsc online", "mpsc exam", "mpsc talathi", "mpsc syllabus", "mpsc portal", "mpsc last date", "mpsc question paper",
            "talathi bharti", "talathi bharti 2026", "talathi form", "talathi apply online",
            # Google Trends Rising Queries (MPSC & Talathi)
            "mahakosh", "mpsc helpline number", "dvet admission", "digital satbara", "rcms", "mahabhumi", "pradhan mantri fasal bima yojana", "mazi naukari.com", "digital 712", "caste validity",
            "mpsc login registration", "maharashtra lokseva aayog", "mahaonline", "talathi online application", "talathi bharti 2026 apply online last date", "talathi bharti last date 2026", "majhi naukri talathi bharti 2026",
            # Original MPSC & Talathi Keywords
            "MPSC", "MPSC 2026", "MPSC recruitment 2026", "MPSC jobs 2026", "MPSC exam 2026", "MPSC latest update", "MPSC notification 2026", "MPSC vacancy 2026",
            "MPSC Rajyaseva", "MPSC Rajyaseva 2026", "MPSC Rajyaseva exam date", "MPSC Rajyaseva syllabus", "MPSC Rajyaseva notification", "MPSC Rajyaseva admit card", "MPSC Rajyaseva result", "MPSC Rajyaseva answer key",
            "MPSC Group C", "MPSC Group C 2026", "MPSC Group C recruitment", "MPSC Group C vacancy", "MPSC Group C syllabus", "MPSC Group C exam date", "MPSC Group C result",
            "Talathi Bharti 2026", "Talathi Bharti", "Talathi recruitment 2026", "Talathi vacancy 2026", "Talathi exam date 2026", "Talathi syllabus 2026", "Talathi application form", "Talathi apply online", "Talathi notification", "Talathi result", "Talathi admit card",
            "तलाठी भरती 2026", "तलाठी भरती", "तलाठी जाहिरात", "तलाठी परीक्षा", "तलाठी अर्ज"
        ]
    },
    "07_UPSC_Universe": {
        "title": "UPSC Keyword Universe",
        "category": "Department/Exam-wise",
        "sub_category": "UPSC Central Services",
        "intent": "Informational/Commercial",
        "target_page": "/jobs?category=upsc",
        "keywords": [
            "UPSC", "UPSC 2026", "UPSC recruitment", "UPSC notification", "UPSC exam date", "UPSC syllabus", "UPSC admit card", "UPSC result", "UPSC answer key", "UPSC current affairs", "UPSC preparation",
            "UPSC CSE", "UPSC Civil Services", "UPSC IAS", "UPSC CDS", "UPSC CAPF", "UPSC EPFO", "UPSC ESE"
        ]
    },
    "08_SSC_Universe": {
        "title": "SSC Keyword Universe",
        "category": "Department/Exam-wise",
        "sub_category": "SSC Central Exams",
        "intent": "Informational/Commercial",
        "target_page": "/jobs?category=ssc",
        "keywords": [
            # Google Trends Top Queries (SSC)
            "ssc cgl", "ssc exam", "ssc gd", "ssc result", "ssc exam date",
            # Google Trends Rising Queries (SSC)
            "how to check ssc 2026 result by sms", "ssc result bd", "education board result", "ssc board login", "ssc board pune", "ssc results 2026",
            # Original SSC Keywords
            "SSC", "SSC 2026", "SSC recruitment 2026", "SSC jobs 2026", "SSC notification", "SSC exam date", "SSC admit card", "SSC result", "SSC answer key",
            "SSC CGL", "SSC CGL 2026", "SSC CHSL", "SSC CHSL 2026", "SSC MTS", "SSC GD", "SSC GD 2026", "SSC JE", "SSC CPO", "SSC selection post"
        ]
    },
    "09_Railway_RRB_Universe": {
        "title": "Railway & RRB Keyword Universe",
        "category": "Department/Exam-wise",
        "sub_category": "Indian Railways / RRB",
        "intent": "Informational/Commercial",
        "target_page": "/jobs?category=railway or /category/railway-jobs",
        "keywords": [
            # Google Trends Top Queries (RRB)
            "rrb", "rrb 2026", "rrb group d", "rrb controller", "rrb admit card", "section controller rrb",
            # Google Trends Rising Queries (RRB)
            "assam career", "ojas", "rrb apprentice recruitment 2026", "railway jobs notification 2026", "tnpsc", "rrb section controller apply online", "rrb group d exam date 2026", "rrb je notification 2026", "rrb recruitment",
            # Original Railway & RRB Keywords
            "railway jobs", "railway jobs 2026", "railway recruitment 2026", "railway vacancy", "railway government jobs", "railway sarkari naukri", "RRB recruitment 2026",
            "RRB NTPC", "RRB NTPC 2026", "RRB NTPC exam date", "RRB NTPC admit card", "RRB NTPC result", "RRB NTPC syllabus", "RRB NTPC answer key", "RRB NTPC vacancy", "RRB NTPC notification",
            "RRB ALP", "RRB ALP 2026", "RRB ALP admit card", "RRB ALP result", "RRB ALP answer key", "RRB ALP exam date",
            "RRB Group D", "RRB Technician", "RRB JE", "RRB Assistant", "Railway Apprentice", "Railway Recruitment Board", "railway apprentice jobs", "railway jobs after 10th", "railway jobs after 12th", "railway jobs for ITI"
        ]
    },
    "10_Banking_Universe": {
        "title": "Banking Keyword Universe (SBI, IBPS, RBI)",
        "category": "Department/Exam-wise",
        "sub_category": "Banking & Financial Sector",
        "intent": "Informational/Commercial",
        "target_page": "/jobs?category=banking or /category/banking-jobs",
        "keywords": [
            # Google Trends Rising Query (Banking)
            "sbi recruitment 2026",
            # Original Banking Keywords
            "SBI jobs", "SBI recruitment 2026", "SBI PO", "SBI Clerk", "SBI PO 2026", "SBI Clerk 2026", "SBI PO syllabus", "SBI PO admit card", "SBI PO result",
            "IBPS", "IBPS recruitment 2026", "IBPS PO", "IBPS Clerk", "IBPS PO 2026", "IBPS Clerk 2026", "IBPS RRB", "IBPS RRB 2026", "IBPS exam date", "IBPS result",
            "RBI recruitment", "RBI jobs 2026", "RBI Grade B", "RBI Grade B 2026", "RBI Assistant", "RBI Assistant 2026", "RBI result", "RBI admit card"
        ]
    },
    "11_Police_Bharti_Universe": {
        "title": "Police Bharti Keyword Universe",
        "category": "Department/Exam-wise",
        "sub_category": "State Police & Uniform Forces",
        "intent": "Informational/Commercial",
        "target_page": "/jobs?category=police or /category/police-jobs",
        "keywords": [
            # Google Trends Top Queries (Police Bharti)
            "police bharti 2026", "up police bharti", "police bharti paper", "mp police bharti", "maharashtra police bharti", "police bharti question paper",
            # Google Trends Rising Queries (Police Bharti)
            "police bharti board", "up police bharti board", "mahajyoti", "gujarat police bharti 2026", "police bharti age limit", "mumbai police bharti",
            # Original Police Bharti Keywords
            "police bharti", "Maharashtra police bharti", "Maharashtra police bharti 2026", "police recruitment 2026", "police vacancy 2026", "police constable recruitment", "police bharti syllabus", "police bharti exam date", "police bharti admit card", "police bharti result", "police bharti physical test", "police bharti form", "police bharti apply online",
            "पोलीस भरती", "पोलीस भरती 2026", "महाराष्ट्र पोलीस भरती"
        ]
    },
    "12_Admit_Cards_Universe": {
        "title": "Admit Card Keyword Universe",
        "category": "Informational/Long-tail",
        "sub_category": "Hall Tickets & Admit Cards",
        "intent": "Transactional/Informational",
        "target_page": "/admit-cards",
        "keywords": [
            # Google Trends Top Queries (Admit Cards)
            "rrb admit card",
            # Original Admit Card Keywords
            "admit card", "government exam admit card", "govt admit card", "sarkari exam admit card", "admit card 2026", "latest admit card", "admit card download", "exam hall ticket", "hall ticket", "MPSC admit card", "SSC admit card", "RRB admit card", "UPSC admit card", "police bharti admit card"
        ]
    },
    "13_Results_Cutoff_MeritList_Universe": {
        "title": "Results, Cutoff & Merit List Keyword Universe",
        "category": "Informational/Long-tail",
        "sub_category": "Exam Results & Scorecards",
        "intent": "Informational/Transactional",
        "target_page": "/exams (Results)",
        "keywords": [
            # Google Trends Top & Rising Queries (Results)
            "ssc result", "how to check ssc 2026 result by sms", "ssc result bd", "education board result", "ssc results 2026",
            # Original Results Keywords
            "government exam result", "govt result", "sarkari result", "government result 2026", "exam result", "latest government result", "MPSC result", "UPSC result", "SSC result", "RRB result", "railway result", "police bharti result", "merit list", "cut off", "cutoff marks", "final result", "result date", "result PDF"
        ]
    },
    "14_Exam_Dates_Calendar_Universe": {
        "title": "Exam Dates & Calendar Keyword Universe",
        "category": "Informational/Long-tail",
        "sub_category": "Schedule & Timetables",
        "intent": "Informational",
        "target_page": "/exam-calendar",
        "keywords": [
            # Google Trends Top & Rising Queries (Exam Dates)
            "ssc exam date", "rrb group d exam date 2026",
            # Original Exam Dates Keywords
            "exam date", "exam date 2026", "government exam date", "govt exam date", "MPSC exam date", "UPSC exam date", "SSC exam date", "RRB exam date", "railway exam date", "police bharti exam date", "admit card date", "answer key date", "result date"
        ]
    },
    "15_Syllabus_Exam_Pattern_Universe": {
        "title": "Syllabus & Exam Pattern Keyword Universe",
        "category": "Informational/Long-tail",
        "sub_category": "Exam Preparation Materials",
        "intent": "Informational",
        "target_page": "/study-material",
        "keywords": [
            # Google Trends Top & Rising Queries (Syllabus & Question Papers)
            "mpsc syllabus", "mpsc question paper", "police bharti paper", "police bharti question paper",
            # Original Syllabus Keywords
            "MPSC syllabus", "UPSC syllabus", "SSC CGL syllabus", "SSC CHSL syllabus", "RRB NTPC syllabus", "RRB Group D syllabus", "police bharti syllabus", "Talathi syllabus", "SBI PO syllabus", "SBI Clerk syllabus", "IBPS PO syllabus", "IBPS Clerk syllabus",
            "MPSC syllabus 2026", "MPSC syllabus PDF", "MPSC syllabus Marathi", "MPSC exam pattern", "MPSC syllabus subject",
            "UPSC syllabus 2026", "UPSC syllabus PDF", "UPSC syllabus Marathi", "UPSC exam pattern",
            "SSC CGL syllabus 2026", "SSC CGL syllabus PDF", "SSC CGL exam pattern",
            "police bharti syllabus 2026", "police bharti syllabus PDF", "police bharti syllabus Marathi", "police bharti exam pattern"
        ]
    },
    "16_Current_Affairs_Daily_Quiz_Universe": {
        "title": "Current Affairs & Daily Quiz Keyword Universe",
        "category": "Informational/Long-tail",
        "sub_category": "General Knowledge & Daily Updates",
        "intent": "Informational",
        "target_page": "/current-affairs",
        "keywords": [
            "current affairs today", "current affairs 2026", "daily current affairs", "current affairs today in Marathi", "current affairs for MPSC", "current affairs for UPSC", "current affairs for SSC", "current affairs PDF", "monthly current affairs", "weekly current affairs", "Maharashtra current affairs", "India current affairs", "current affairs quiz", "current affairs questions"
        ]
    }
}

# 2. Read existing Excel dataset
wb_existing = openpyxl.load_workbook(excel_path)
ws_keywords = wb_existing["Keywords"]

existing_entries = []
seen_keywords_lower = set()

for r in range(2, ws_keywords.max_row + 1):
    s_no = ws_keywords.cell(r, 1).value
    kw = ws_keywords.cell(r, 2).value
    cat = ws_keywords.cell(r, 3).value
    sub_cat = ws_keywords.cell(r, 4).value
    intent = ws_keywords.cell(r, 5).value
    
    if kw and str(kw).strip():
        kw_clean = str(kw).strip()
        kw_lower = kw_clean.lower()
        if kw_lower not in seen_keywords_lower:
            seen_keywords_lower.add(kw_lower)
            existing_entries.append({
                "keyword": kw_clean,
                "category": cat or "General",
                "sub_category": sub_cat or "General",
                "intent": intent or "Informational"
            })

print(f"Loaded {len(existing_entries)} unique keywords from existing Excel sheet.")

# 3. Process and Merge Prompt Keywords
cluster_data = {}
new_added_count = 0

for file_prefix, info in prompt_clusters.items():
    cluster_data[file_prefix] = {
        "title": info["title"],
        "category": info["category"],
        "sub_category": info["sub_category"],
        "intent": info["intent"],
        "target_page": info["target_page"],
        "entries": []
    }
    for kw in info["keywords"]:
        kw_clean = kw.strip()
        kw_lower = kw_clean.lower()
        
        # Add to cluster file entries
        cluster_data[file_prefix]["entries"].append({
            "keyword": kw_clean,
            "category": info["category"],
            "sub_category": info["sub_category"],
            "intent": info["intent"],
            "target_page": info["target_page"]
        })
        
        # Add to master excel entries if not seen
        if kw_lower not in seen_keywords_lower:
            seen_keywords_lower.add(kw_lower)
            existing_entries.append({
                "keyword": kw_clean,
                "category": info["category"],
                "sub_category": info["sub_category"],
                "intent": info["intent"]
            })
            new_added_count += 1

print(f"Added {new_added_count} new unique keywords from prompt frameworks.")
print(f"Total Master Keyword Count: {len(existing_entries)}")

# Map existing keywords into appropriate prompt clusters if they fit
for entry in existing_entries:
    kw_l = entry["keyword"].lower()
    cat_l = str(entry["category"]).lower()
    sub_l = str(entry["sub_category"]).lower()
    
    assigned = False
    
    # Categorize into cluster files for comprehensive output files
    if "10th" in kw_l or "12th" in kw_l or "iti" in kw_l or "diploma" in kw_l or "graduate" in kw_l or "engineering" in kw_l or "qualification" in cat_l:
        cluster_data["04_Qualification_Universe"]["entries"].append(entry)
        assigned = True
    elif "maharashtra" in kw_l or "pune" in kw_l or "mumbai" in kw_l or "nagpur" in kw_l or "nashik" in kw_l or "district" in cat_l:
        cluster_data["05_Maharashtra_Districts_Universe"]["entries"].append(entry)
        assigned = True
    elif "mpsc" in kw_l or "talathi" in kw_l or "rajyaseva" in kw_l or "group c" in kw_l:
        cluster_data["06_MPSC_Talathi_GroupC_Universe"]["entries"].append(entry)
        assigned = True
    elif "upsc" in kw_l or "ias" in kw_l or "cds" in kw_l or "epfo" in kw_l:
        cluster_data["07_UPSC_Universe"]["entries"].append(entry)
        assigned = True
    elif "ssc" in kw_l or "cgl" in kw_l or "chsl" in kw_l or "mts" in kw_l:
        cluster_data["08_SSC_Universe"]["entries"].append(entry)
        assigned = True
    elif "railway" in kw_l or "rrb" in kw_l or "ntpc" in kw_l or "alp" in kw_l:
        cluster_data["09_Railway_RRB_Universe"]["entries"].append(entry)
        assigned = True
    elif "sbi" in kw_l or "ibps" in kw_l or "rbi" in kw_l or "bank" in kw_l:
        cluster_data["10_Banking_Universe"]["entries"].append(entry)
        assigned = True
    elif "police" in kw_l or "constable" in kw_l or "पोलीस" in kw_l:
        cluster_data["11_Police_Bharti_Universe"]["entries"].append(entry)
        assigned = True
    elif "admit card" in kw_l or "hall ticket" in kw_l:
        cluster_data["12_Admit_Cards_Universe"]["entries"].append(entry)
        assigned = True
    elif "result" in kw_l or "cutoff" in kw_l or "cut off" in kw_l or "merit" in kw_l:
        cluster_data["13_Results_Cutoff_MeritList_Universe"]["entries"].append(entry)
        assigned = True
    elif "exam date" in kw_l or "calendar" in kw_l or "timetable" in kw_l:
        cluster_data["14_Exam_Dates_Calendar_Universe"]["entries"].append(entry)
        assigned = True
    elif "syllabus" in kw_l or "pattern" in kw_l:
        cluster_data["15_Syllabus_Exam_Pattern_Universe"]["entries"].append(entry)
        assigned = True
    elif "current affairs" in kw_l or "quiz" in kw_l or "gk" in kw_l:
        cluster_data["16_Current_Affairs_Daily_Quiz_Universe"]["entries"].append(entry)
        assigned = True
    elif "last date" in kw_l or "closing" in kw_l or "ending" in kw_l or "apply online" in kw_l:
        cluster_data["03_Deadline_Closing_Soon"]["entries"].append(entry)
        assigned = True
    elif "latest" in kw_l or "new" in kw_l or "upcoming" in kw_l:
        cluster_data["02_Latest_Jobs"]["entries"].append(entry)
        assigned = True

# Deduplicate individual cluster entries by keyword string
for file_prefix, cdata in cluster_data.items():
    unique_cluster_entries = []
    seen_cluster_kw = set()
    for e in cdata["entries"]:
        kw_norm = e["keyword"].strip().lower()
        if kw_norm not in seen_cluster_kw:
            seen_cluster_kw.add(kw_norm)
            unique_cluster_entries.append(e)
    cdata["entries"] = unique_cluster_entries

# 4. Generate Individual Category Markdown and Text Files
for file_prefix, cdata in cluster_data.items():
    md_filepath = os.path.join(base_dir, f"{file_prefix}.md")
    txt_filepath = os.path.join(base_dir, f"{file_prefix}.txt")
    
    # Write MD
    md_lines = []
    md_lines.append(f"# {cdata['title']}")
    md_lines.append("")
    md_lines.append(f"**Primary Category:** `{cdata['category']}`  ")
    md_lines.append(f"**Sub-Category:** `{cdata['sub_category']}`  ")
    md_lines.append(f"**Search Intent:** `{cdata['intent']}`  ")
    md_lines.append(f"**Recommended Target Landing Page:** `{cdata['target_page']}`  ")
    md_lines.append(f"**Total Keywords in Cluster:** **{len(cdata['entries'])}**  ")
    md_lines.append("")
    md_lines.append("---")
    md_lines.append("")
    md_lines.append("## 📋 Keyword Table & Intent Mapping")
    md_lines.append("")
    md_lines.append("| # | Keyword | Category | Sub-Category | Search Intent |")
    md_lines.append("|---|---|---|---|---|")
    
    for idx, e in enumerate(cdata["entries"], 1):
        cat_disp = e.get("category", cdata["category"])
        sub_disp = e.get("sub_category", cdata["sub_category"])
        intent_disp = e.get("intent", cdata["intent"])
        md_lines.append(f"| {idx} | `{e['keyword']}` | {cat_disp} | {sub_disp} | {intent_disp} |")
    
    md_lines.append("")
    md_lines.append("---")
    md_lines.append("")
    md_lines.append("## 📄 Plain Text List")
    md_lines.append("")
    md_lines.append("```text")
    for e in cdata["entries"]:
        md_lines.append(e["keyword"])
    md_lines.append("```")
    
    with open(md_filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines))
        
    # Write TXT
    txt_lines = [f"{cdata['title']} - Total Keywords: {len(cdata['entries'])}", "=" * 70, ""]
    for e in cdata["entries"]:
        txt_lines.append(e["keyword"])
        
    with open(txt_filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(txt_lines))

# 5. Generate Master Directory Overview Document
master_md_filepath = os.path.join(base_dir, "00_MASTER_KEYWORD_DIRECTORY.md")
master_txt_filepath = os.path.join(base_dir, "00_MASTER_KEYWORD_DIRECTORY.txt")

master_md = []
master_md.append("# 🔑 SearchSarkariNaukri.com — Master Keyword Research & Cluster Directory")
master_md.append("")
master_md.append("**Audit & Optimization Date:** 2026-08-10  ")
master_md.append(f"**Total Master Unique Keywords:** **{len(existing_entries)}**  ")
master_md.append(f"**Total Master Clusters:** **{len(prompt_clusters)}**  ")
master_md.append("")
master_md.append("---")
master_md.append("")
master_md.append("## 💡 SEO Strategy & Keyword Placement Rules")
master_md.append("> **Rule:** Do NOT put all keyword types on a single page. Follow the **1 Keyword Cluster → 1 Search Intent → 1 Dedicated Page** model.")
master_md.append("")
master_md.append("---")
master_md.append("")
master_md.append("## 📊 Master Clusters Overview")
master_md.append("")
master_md.append("| # | Cluster Name | Primary Target Page | Intent Type | Unique Keywords | Files |")
master_md.append("|---|---|---|---|---|---|")

base_dir_url = base_dir.replace('\\', '/')
for idx, (prefix, cdata) in enumerate(cluster_data.items(), 1):
    count = len(cdata["entries"])
    master_md.append(f"| {idx} | **{cdata['title']}** | `{cdata['target_page']}` | {cdata['intent']} | **{count}** | [`{prefix}.md`](file:///{base_dir_url}/{prefix}.md) \| [`txt`](file:///{base_dir_url}/{prefix}.txt) |")

master_md.append(f"| | **TOTAL UNIQUE KEYWORDS** | **All Clusters Combined** | | **{len(existing_entries)}** | |")
master_md.append("")
master_md.append("---")
master_md.append("")
master_md.append("## 📁 Full Master Keyword Index")
master_md.append("")
master_md.append("```text")
for idx, entry in enumerate(existing_entries, 1):
    master_md.append(f"{idx}. {entry['keyword']}")
master_md.append("```")

with open(master_md_filepath, "w", encoding="utf-8") as f:
    f.write("\n".join(master_md))

# Master TXT
master_txt = [f"SearchSarkariNaukri.com Master Keyword Database - Total: {len(existing_entries)} Keywords", "=" * 80, ""]
for idx, entry in enumerate(existing_entries, 1):
    master_txt.append(entry["keyword"])

with open(master_txt_filepath, "w", encoding="utf-8") as f:
    f.write("\n".join(master_txt))

# 6. Update Excel Spreadsheet
print("Updating Excel file with complete expanded dataset...")
wb_out = openpyxl.Workbook()

# Summary sheet
ws_sum = wb_out.active
ws_sum.title = "Summary"
ws_sum.append(["SearchSarkariNaukri.com — Master Keyword Research Framework", None])
ws_sum.append([f"Total unique keywords: {len(existing_entries)}", None])
ws_sum.append([])
ws_sum.append(["Category / Cluster Name", "Target Page", "Keyword Count"])

category_counts = {}
for entry in existing_entries:
    cat = entry.get("category", "General")
    category_counts[cat] = category_counts.get(cat, 0) + 1

for cat, count in category_counts.items():
    ws_sum.append([cat, "See Details in Keywords Tab", count])

ws_sum.append(["TOTAL UNIQUE KEYWORDS", "Master Dataset", len(existing_entries)])

# Keywords sheet
ws_kw = wb_out.create_sheet(title="Keywords")
ws_kw.append(["S.No", "Keyword", "Category", "Sub-Category", "Search Intent"])

for idx, entry in enumerate(existing_entries, 1):
    ws_kw.append([idx, entry["keyword"], entry["category"], entry["sub_category"], entry["intent"]])

wb_out.save(excel_path)
wb_out.save(excel_backup_path)

print(f"Successfully updated original Excel: {excel_path}")
print(f"Saved master copy to: {excel_backup_path}")
print(f"Generated {len(cluster_data)} individual category files in {base_dir}")
